"""
Porsche Parts Catalog PDF → Excel Converter
=============================================
A small desktop app that extracts parts tables from Porsche-style
catalog PDFs and exports them to a formatted .xlsx file.

Dependencies:
    pip install PyMuPDF openpyxl tkinterdnd2

Run:
    python pdf_to_excel.py
"""

# ── Standard library ──────────────────────────────────────────────
import os
import re
import sys
import subprocess
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

# ── Dependency check ──────────────────────────────────────────────
def check_deps():
    """Return list of missing package names."""
    missing = []
    try:
        import fitz  # noqa: F401 – PyMuPDF
    except ImportError:
        missing.append("PyMuPDF")
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        missing.append("openpyxl")
    try:
        import tkinterdnd2  # noqa: F401
    except ImportError:
        missing.append("tkinterdnd2")
    return missing

missing = check_deps()
if missing:
    msg = (
        "Missing packages: " + ", ".join(missing) + "\n\n"
        "Install them with:\n"
        "  pip install PyMuPDF openpyxl tkinterdnd2\n"
    )
    # Try a Tk messagebox first; fall back to console
    try:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Missing Dependencies", msg)
        root.destroy()
    except Exception:
        pass
    print(msg, file=sys.stderr)
    sys.exit(1)

import fitz                       # PDF text extraction (PyMuPDF)
import openpyxl                   # Excel creation
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from tkinterdnd2 import TkinterDnD, DND_FILES  # native drag-and-drop


# ══════════════════════════════════════════════════════════════════
#  PDF PARSING
# ══════════════════════════════════════════════════════════════════
#
# The PDF text (via PyMuPDF) comes out with each field on its own line.
#
# Single-digit item (1–9):          Multi-digit item (10+):
#   1                                 10 TORX SCREW BM6X25
#   Z CRANKCASE                       14 PAF008127
#   1
#   0PB101040B
#
# With M marker:                    With M marker (10+):
#   3                                 26 STUD M12X1.5X43
#   SCREW PLUG M20X1.5X18            X
#   X                                 6
#   2                                 WHT009844
#   WHT008323
#
# The parser uses a state machine:
#   SCAN       → looking for next item number
#   WAIT_DESC  → single-digit item found, next line is description
#   WAIT_QTY   → description known, next is "X" or qty [+ material]
#   WAIT_QTY2  → got "X" marker, next is qty [+ material]
#   WAIT_MAT   → got qty alone on a line, next line is material

# Lines to skip in page headers
SKIP_LINES = {"MOTORSPORT", "911 GT3 Cup"}

# Category line:  "1.2  crank case" or "10.3 greases and liquides"
CATEGORY_RE = re.compile(r"^(\d+\.\d+)\s+(.+)$")

# Table column header
TABLE_HDR_RE = re.compile(r"^Item no\.\s+Description$", re.IGNORECASE)
TABLE_HDR2_RE = re.compile(r"^M\s+Pc\.\s+Material$", re.IGNORECASE)

# A line that is just a single digit (1–9) → single-digit item number
SINGLE_ITEM_RE = re.compile(r"^(\d)$")

# A line starting with 2+ digits then a space and text → multi-digit item
MULTI_ITEM_RE = re.compile(r"^(\d{2,})\s+(.+)$")

# A line with a number followed by text → qty + material on same line
QTY_MAT_RE = re.compile(r"^(\d+)\s+(.+)$")

# A line that is just digits → qty alone (or page number, handled by context)
QTY_ONLY_RE = re.compile(r"^(\d+)$")

# "X" marker + qty + material all on one line:  "X 32 N 10879301"
X_QTY_MAT_RE = re.compile(r"^X\s+(\d+)\s+(.+)$")

# "X" marker + qty on one line (no material):  "X 32"
X_QTY_ONLY_RE = re.compile(r"^X\s+(\d+)$")


def parse_pdf(pdf_path, log_fn=None):
    """
    Parse *pdf_path* and return a list of dicts, one per parts row.

    Each dict has keys:
        Category, Item no, Description, M, Pc., Material, Page
    """
    if log_fn is None:
        log_fn = lambda msg: None  # noqa: E731

    doc = fitz.open(pdf_path)
    log_fn(f"Opened PDF: {os.path.basename(pdf_path)}  ({len(doc)} pages)")

    rows = []
    current_category = ""

    for page_num in range(len(doc)):
        page = doc[page_num]
        text = page.get_text("text")
        lines = text.split("\n")

        in_data = False    # True after we see the table header
        state = "SCAN"
        item_no = ""
        description = ""
        m_marker = ""
        qty = ""

        def _save_row(material=""):
            nonlocal item_no, description, m_marker, qty, state
            rows.append({
                "Category":    current_category,
                "Item no":     item_no,
                "Description": description,
                "M":           m_marker,
                "Pc.":         qty,
                "Material":    material.strip(),
                "Page":        page_num + 1,
            })
            item_no = description = m_marker = qty = ""
            state = "SCAN"

        for line in lines:
            line = line.strip()
            if not line:
                continue

            # ── Skip fixed header lines ───────────────────────
            if line in SKIP_LINES:
                continue
            if line.startswith("Creation date:") or line.startswith("Model "):
                continue

            # ── Page number at top (before data section) ──────
            if not in_data and line == str(page_num + 1):
                continue

            # ── Section-only pages like "1  engine" ───────────
            # These are chapter dividers with no table; skip them
            if not in_data and re.match(r"^\d+\s+\w+$", line) and not CATEGORY_RE.match(line):
                continue

            # ── Category line (e.g. "1.3  crank shaft") ──────
            cat = CATEGORY_RE.match(line)
            if cat and not in_data:
                current_category = line.strip()
                log_fn(f"  Page {page_num + 1}: category → {current_category}")
                continue

            # ── Table header lines ────────────────────────────
            if TABLE_HDR_RE.match(line):
                continue
            if TABLE_HDR2_RE.match(line):
                in_data = True
                state = "SCAN"
                continue

            if not in_data:
                continue

            # ══ State machine for parsing rows ════════════════

            if state == "SCAN":
                # Single-digit item number alone on a line
                sm = SINGLE_ITEM_RE.match(line)
                if sm:
                    item_no = sm.group(1)
                    state = "WAIT_DESC"
                    continue
                # Multi-digit item number + description
                mm = MULTI_ITEM_RE.match(line)
                if mm:
                    item_no = mm.group(1)
                    description = mm.group(2)
                    state = "WAIT_QTY"
                    continue

            elif state == "WAIT_DESC":
                # This line is the description text
                description = line
                state = "WAIT_QTY"
                continue

            elif state == "WAIT_QTY":
                # "X qty material" all on one line (e.g. "X 32 N 10879301")
                xqm = X_QTY_MAT_RE.match(line)
                if xqm:
                    m_marker = "X"
                    qty = xqm.group(1)
                    _save_row(xqm.group(2))
                    continue
                # "X qty" on one line (e.g. "X 32")
                xqo = X_QTY_ONLY_RE.match(line)
                if xqo:
                    m_marker = "X"
                    qty = xqo.group(1)
                    state = "WAIT_MAT"
                    continue
                # "X" alone → M marker
                if line == "X":
                    m_marker = "X"
                    state = "WAIT_QTY2"
                    continue
                # qty + material on same line
                qm = QTY_MAT_RE.match(line)
                if qm:
                    qty = qm.group(1)
                    _save_row(qm.group(2))
                    continue
                # qty alone
                qo = QTY_ONLY_RE.match(line)
                if qo:
                    qty = qo.group(1)
                    state = "WAIT_MAT"
                    continue
                # Fallback: unrecognized line — reset to prevent sync loss
                state = "SCAN"

            elif state == "WAIT_QTY2":
                # After "X" marker — expect qty [+ material]
                qm = QTY_MAT_RE.match(line)
                if qm:
                    qty = qm.group(1)
                    _save_row(qm.group(2))
                    continue
                qo = QTY_ONLY_RE.match(line)
                if qo:
                    qty = qo.group(1)
                    state = "WAIT_MAT"
                    continue
                # Fallback: reset
                state = "SCAN"

            elif state == "WAIT_MAT":
                # This line is the material value
                _save_row(line)
                continue

    doc.close()
    log_fn(f"Extracted {len(rows)} rows total.")
    return rows


# ══════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════

COLUMNS = ["Category", "Item no", "Description", "M", "Pc.", "Material", "Page"]
COL_WIDTHS = {
    "Category":    28,
    "Item no":     10,
    "Description": 40,
    "M":            5,
    "Pc.":          6,
    "Material":    20,
    "Page":         7,
}


def export_to_excel(rows, save_path, log_fn=None):
    """Write *rows* (list of dicts) to an Excel file at *save_path*."""
    if log_fn is None:
        log_fn = lambda msg: None  # noqa: E731

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parts"

    # ── Header row ────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # ── Data rows ─────────────────────────────────────────────────
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))

    # ── Column widths ─────────────────────────────────────────────
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_name]

    # ── Freeze top row ────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Excel table (structured reference) ────────────────────────
    last_col_letter = get_column_letter(len(COLUMNS))
    last_row = len(rows) + 1  # +1 for header
    table_ref = f"A1:{last_col_letter}{last_row}"
    table = Table(displayName="PartsTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    wb.save(save_path)
    log_fn(f"Saved Excel → {save_path}")


# ══════════════════════════════════════════════════════════════════
#  GUI
# ══════════════════════════════════════════════════════════════════

class App(TkinterDnD.Tk):
    """Main application window."""

    # ── Color palette ─────────────────────────────────────────────
    BG           = "#1a1a2e"    # deep navy background
    BG_CARD      = "#16213e"    # card/panel background
    BG_SURFACE   = "#0f3460"    # elevated surface
    ACCENT       = "#e94560"    # Porsche-red accent
    ACCENT_HOVER = "#ff6b81"
    TEXT         = "#eaeaea"    # primary text
    TEXT_DIM     = "#8892a4"    # secondary text
    TEXT_ACCENT  = "#53c0f0"    # status / link color
    BORDER       = "#2a2a4a"    # subtle borders
    DROP_BG      = "#1c2541"    # drop zone fill
    DROP_BORDER  = "#e94560"    # drop zone dashed border
    LOG_BG       = "#0d1b2a"    # log area background
    SUCCESS      = "#00d26a"
    WARNING      = "#ffbe0b"

    def __init__(self):
        super().__init__()
        self.title("Porsche PDF \u2192 Excel Converter")
        self.resizable(False, False)
        self.configure(bg=self.BG)

        # State
        self.pdf_path = None
        self.rows = []
        self.open_after = tk.BooleanVar(value=False)

        self._configure_styles()
        self._build_ui()
        self._center_window(560, 560)

    # ── ttk Styles ────────────────────────────────────────────────

    def _configure_styles(self):
        style = ttk.Style(self)
        style.theme_use("clam")

        # Primary button (accent)
        style.configure("Accent.TButton",
            font=("Segoe UI Semibold", 10),
            foreground="#ffffff",
            background=self.ACCENT,
            borderwidth=0,
            padding=(16, 8),
            focuscolor="none",
        )
        style.map("Accent.TButton",
            background=[("active", self.ACCENT_HOVER), ("disabled", "#3a3a5a")],
            foreground=[("disabled", "#666")],
        )

        # Secondary button
        style.configure("Secondary.TButton",
            font=("Segoe UI", 10),
            foreground=self.TEXT,
            background=self.BG_SURFACE,
            borderwidth=0,
            padding=(14, 8),
            focuscolor="none",
        )
        style.map("Secondary.TButton",
            background=[("active", "#1a4a80"), ("disabled", "#2a2a4a")],
            foreground=[("disabled", "#555")],
        )

        # Progress bar
        style.configure("Custom.Horizontal.TProgressbar",
            troughcolor=self.BG_CARD,
            background=self.ACCENT,
            borderwidth=0,
            thickness=4,
        )

        # Checkbox
        style.configure("Dark.TCheckbutton",
            font=("Segoe UI", 9),
            foreground=self.TEXT_DIM,
            background=self.BG,
            focuscolor="none",
        )
        style.map("Dark.TCheckbutton",
            background=[("active", self.BG)],
        )

    # ── Layout ────────────────────────────────────────────────────

    def _build_ui(self):
        # ── Header ────────────────────────────────────────────────
        header = tk.Frame(self, bg=self.BG_CARD, height=64)
        header.pack(fill="x")
        header.pack_propagate(False)

        title_frame = tk.Frame(header, bg=self.BG_CARD)
        title_frame.pack(expand=True)

        # Red accent bar on the left of the title
        tk.Frame(title_frame, bg=self.ACCENT, width=4, height=32).pack(
            side="left", padx=(0, 10), pady=16
        )
        tk.Label(
            title_frame,
            text="PORSCHE",
            bg=self.BG_CARD, fg=self.TEXT,
            font=("Segoe UI Black", 16),
        ).pack(side="left")
        tk.Label(
            title_frame,
            text="  PDF \u2192 Excel",
            bg=self.BG_CARD, fg=self.TEXT_DIM,
            font=("Segoe UI Light", 16),
        ).pack(side="left")

        # Thin accent line under header
        tk.Frame(self, bg=self.ACCENT, height=2).pack(fill="x")

        # ── Main content area ─────────────────────────────────────
        content = tk.Frame(self, bg=self.BG)
        content.pack(fill="both", expand=True, padx=20, pady=16)

        # ── Drop zone ────────────────────────────────────────────
        # Outer border frame to simulate a dashed/colored border
        drop_border = tk.Frame(content, bg=self.DROP_BORDER, padx=2, pady=2)
        drop_border.pack(fill="x", pady=(0, 12))

        self.drop_frame = tk.Frame(drop_border, bg=self.DROP_BG, height=90)
        self.drop_frame.pack(fill="x")
        self.drop_frame.pack_propagate(False)

        self.drop_icon = tk.Label(
            self.drop_frame, text="\u2193",
            bg=self.DROP_BG, fg=self.ACCENT,
            font=("Segoe UI", 22),
        )
        self.drop_icon.pack(pady=(14, 0))

        self.drop_label = tk.Label(
            self.drop_frame,
            text="Drag & drop a PDF here",
            bg=self.DROP_BG, fg=self.TEXT_DIM,
            font=("Segoe UI", 11),
        )
        self.drop_label.pack(pady=(2, 14))

        # Register drag-and-drop on frame, icon, and label
        for widget in (self.drop_frame, self.drop_icon, self.drop_label):
            widget.drop_target_register(DND_FILES)
            widget.dnd_bind("<<Drop>>", self._on_drop)

        # ── File info ─────────────────────────────────────────────
        self.file_var = tk.StringVar(value="No file selected")
        tk.Label(
            content, textvariable=self.file_var,
            bg=self.BG, fg=self.TEXT_DIM, font=("Segoe UI", 9),
            anchor="w",
        ).pack(fill="x", pady=(0, 10))

        # ── Button row ───────────────────────────────────────────
        btn_frame = tk.Frame(content, bg=self.BG)
        btn_frame.pack(fill="x", pady=(0, 8))

        self.btn_browse = ttk.Button(
            btn_frame, text="Browse PDF",
            style="Secondary.TButton", command=self._browse,
        )
        self.btn_browse.pack(side="left", padx=(0, 6))

        self.btn_convert = ttk.Button(
            btn_frame, text="\u25B6  Convert",
            style="Accent.TButton", command=self._convert,
        )
        self.btn_convert.pack(side="left", padx=6)
        self.btn_convert.state(["disabled"])

        self.btn_save = ttk.Button(
            btn_frame, text="Save As\u2026",
            style="Accent.TButton", command=self._save_as,
        )
        self.btn_save.pack(side="left", padx=6)
        self.btn_save.state(["disabled"])

        self.btn_clear = ttk.Button(
            btn_frame, text="Clear",
            style="Secondary.TButton", command=self._clear,
        )
        self.btn_clear.pack(side="right")

        # ── Options row ──────────────────────────────────────────
        ttk.Checkbutton(
            content, text="Open Excel file after export",
            variable=self.open_after, style="Dark.TCheckbutton",
        ).pack(anchor="w", pady=(0, 8))

        # ── Progress bar ─────────────────────────────────────────
        self.progress = ttk.Progressbar(
            content, mode="indeterminate",
            style="Custom.Horizontal.TProgressbar", length=520,
        )
        self.progress.pack(fill="x", pady=(0, 6))

        # ── Status label ─────────────────────────────────────────
        self.status_var = tk.StringVar(value="\u25CF  Ready")
        self.status_label = tk.Label(
            content, textvariable=self.status_var,
            bg=self.BG, fg=self.TEXT_ACCENT,
            font=("Segoe UI Semibold", 9),
            anchor="w",
        )
        self.status_label.pack(fill="x", pady=(0, 8))

        # ── Log section ──────────────────────────────────────────
        tk.Label(
            content, text="LOG",
            bg=self.BG, fg=self.TEXT_DIM,
            font=("Segoe UI Semibold", 8),
            anchor="w",
        ).pack(fill="x")

        log_frame = tk.Frame(content, bg=self.BORDER, padx=1, pady=1)
        log_frame.pack(fill="both", expand=True, pady=(2, 0))

        self.log_box = tk.Text(
            log_frame, height=10, wrap="word",
            font=("Cascadia Code", 9), bg=self.LOG_BG, fg="#7faacc",
            insertbackground=self.TEXT,
            state="disabled", relief="flat", bd=0,
            padx=10, pady=8,
            selectbackground=self.BG_SURFACE,
        )
        scrollbar = ttk.Scrollbar(log_frame, command=self.log_box.yview)
        self.log_box.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.log_box.pack(side="left", fill="both", expand=True)

    # ── Helpers ───────────────────────────────────────────────────

    def _center_window(self, w, h):
        x = (self.winfo_screenwidth() - w) // 2
        y = (self.winfo_screenheight() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    def _log(self, msg):
        """Append a line to the log box (thread-safe via after)."""
        def _append():
            self.log_box.configure(state="normal")
            self.log_box.insert("end", msg + "\n")
            self.log_box.see("end")
            self.log_box.configure(state="disabled")
        self.after(0, _append)

    def _set_status(self, msg, color=None):
        if color is None:
            color = self.TEXT_ACCENT
        def _update():
            self.status_var.set(msg)
            self.status_label.configure(fg=color)
        self.after(0, _update)

    # ── Drag & drop ──────────────────────────────────────────────

    def _on_drop(self, event):
        """Handle a file dropped onto the drop zone."""
        path = event.data.strip()
        # tkdnd wraps paths with spaces in braces: {C:/my path/file.pdf}
        if path.startswith("{") and path.endswith("}"):
            path = path[1:-1]
        if path.lower().endswith(".pdf"):
            self._load_pdf(path)
        else:
            self._log("Dropped file is not a PDF.")
            self._set_status("\u25CF  Please drop a .pdf file.", self.WARNING)

    # ── Actions ───────────────────────────────────────────────────

    def _browse(self):
        path = filedialog.askopenfilename(
            title="Select a PDF",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
        )
        if path:
            self._load_pdf(path)

    def _load_pdf(self, path):
        self.pdf_path = path
        name = os.path.basename(path)
        self.file_var.set(name)
        self.drop_label.config(text=name, fg=self.TEXT)
        self.drop_icon.config(text="\u2713", fg=self.SUCCESS)
        self.btn_convert.state(["!disabled"])
        self.btn_save.state(["disabled"])
        self.rows = []
        self._log(f"Loaded: {name}")
        self._set_status("\u25CF  PDF loaded \u2013 click Convert", self.SUCCESS)

    def _convert(self):
        """Parse the PDF in a background thread so the UI stays responsive."""
        if not self.pdf_path:
            return

        self.btn_convert.state(["disabled"])
        self.btn_browse.state(["disabled"])
        self.progress.start(12)
        self._set_status("\u25CF  Parsing PDF\u2026", self.WARNING)

        def work():
            try:
                self.rows = parse_pdf(self.pdf_path, log_fn=self._log)
                if not self.rows:
                    self._set_status("\u25CF  No parts rows found", self.WARNING)
                    self._log("No data rows matched. Check the PDF layout.")
                else:
                    self._set_status(f"\u25CF  Done \u2013 {len(self.rows)} rows extracted", self.SUCCESS)
                    self.after(0, lambda: self.btn_save.state(["!disabled"]))
            except Exception as exc:
                self._log(f"ERROR: {exc}")
                self._set_status("\u25CF  Conversion failed \u2013 see log", self.ACCENT)
            finally:
                self.after(0, lambda: self.progress.stop())
                self.after(0, lambda: self.btn_convert.state(["!disabled"]))
                self.after(0, lambda: self.btn_browse.state(["!disabled"]))

        threading.Thread(target=work, daemon=True).start()

    def _save_as(self):
        if not self.rows:
            messagebox.showwarning("Nothing to save", "Convert a PDF first.")
            return

        default_name = os.path.splitext(os.path.basename(self.pdf_path))[0] + ".xlsx"
        path = filedialog.asksaveasfilename(
            title="Save Excel file",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not path:
            return

        try:
            export_to_excel(self.rows, path, log_fn=self._log)
            self._set_status(f"\u25CF  Saved: {os.path.basename(path)}", self.SUCCESS)
            messagebox.showinfo("Success", f"Excel file saved:\n{path}")

            if self.open_after.get():
                # Open with the system's default app
                if sys.platform == "win32":
                    os.startfile(path)
                elif sys.platform == "darwin":
                    subprocess.Popen(["open", path])
                else:
                    subprocess.Popen(["xdg-open", path])

        except PermissionError:
            self._log("ERROR: File is open in another program. Close it and retry.")
            self._set_status("\u25CF  Save failed \u2013 file locked", self.ACCENT)
        except Exception as exc:
            self._log(f"ERROR saving: {exc}")
            self._set_status("\u25CF  Save failed \u2013 see log", self.ACCENT)

    def _clear(self):
        self.pdf_path = None
        self.rows = []
        self.file_var.set("No file selected")
        self.drop_label.config(text="Drag & drop a PDF here", fg=self.TEXT_DIM)
        self.drop_icon.config(text="\u2193", fg=self.ACCENT)
        self.btn_convert.state(["disabled"])
        self.btn_save.state(["disabled"])
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")
        self._set_status("\u25CF  Ready", self.TEXT_ACCENT)


# ══════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    app = App()
    app.mainloop()
